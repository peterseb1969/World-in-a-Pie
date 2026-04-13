"""CSV/XLSX file import service for bulk document creation."""

import csv
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)


class ImportService:
    """Handles parsing CSV/XLSX files and creating documents from rows."""

    @staticmethod
    def detect_format(filename: str) -> str:
        """Detect file format from extension. Returns 'csv' or 'xlsx'."""
        lower = filename.lower()
        if lower.endswith('.xlsx'):
            return 'xlsx'
        if lower.endswith('.xls'):
            return 'xls'
        return 'csv'

    @staticmethod
    def preview(file_content: bytes, filename: str) -> dict[str, Any]:
        """Parse file and return headers + sample rows.

        Returns:
            {format, headers, sample_rows (first 5), total_rows}
        """
        fmt = ImportService.detect_format(filename)

        if fmt == 'xlsx':
            try:
                import openpyxl
            except ImportError:
                raise ValueError("openpyxl not installed. Add it to requirements.txt.")

            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not rows:
                raise ValueError("Empty spreadsheet")

            headers = [str(h) if h is not None else f"column_{i}" for i, h in enumerate(rows[0])]
            data_rows = []
            for row in rows[1:]:
                data_rows.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row) if i < len(headers)})

            return {
                "format": "xlsx",
                "headers": headers,
                "sample_rows": data_rows[:5],
                "total_rows": len(data_rows),
            }

        elif fmt == 'xls':
            raise ValueError("Legacy .xls format not supported. Please save as .xlsx or .csv.")

        else:  # csv
            text = file_content.decode('utf-8-sig')  # Handle BOM
            reader = csv.DictReader(io.StringIO(text))
            headers = reader.fieldnames or []
            if not headers:
                raise ValueError("CSV file has no headers")

            data_rows = []
            for row in reader:
                data_rows.append(dict(row))

            return {
                "format": "csv",
                "headers": list(headers),
                "sample_rows": data_rows[:5],
                "total_rows": len(data_rows),
            }

    @staticmethod
    def parse_rows(file_content: bytes, filename: str) -> tuple[list[str], list[dict[str, str]]]:
        """Parse all rows from file. Returns (headers, rows)."""
        fmt = ImportService.detect_format(filename)

        if fmt == 'xlsx':
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not all_rows:
                raise ValueError("Empty spreadsheet")

            headers = [str(h) if h is not None else f"column_{i}" for i, h in enumerate(all_rows[0])]
            data_rows = []
            for row in all_rows[1:]:
                data_rows.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row) if i < len(headers)})
            return headers, data_rows

        else:  # csv
            text = file_content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            headers = list(reader.fieldnames or [])
            data_rows = [dict(row) for row in reader]
            return headers, data_rows

    @staticmethod
    def build_documents(
        rows: list[dict[str, str]],
        template_id: str,
        column_mapping: dict[str, str],
        namespace: str,
    ) -> list[dict[str, Any]]:
        """Convert parsed rows to document create requests using column mapping.

        Args:
            rows: List of {csv_column: value} dicts
            template_id: Template ID to create documents for
            column_mapping: {csv_column_name: template_field_name}
            namespace: Target namespace

        Returns:
            List of document create request dicts
        """
        documents = []
        for row in rows:
            data = {}
            for csv_col, field_name in column_mapping.items():
                if row.get(csv_col):
                    value = row[csv_col]
                    # Try to convert numeric strings
                    # Leave as string if it fails - let doc-store validation handle it
                    data[field_name] = value

            documents.append({
                "template_id": template_id,
                "namespace": namespace,
                "data": data,
            })

        return documents
